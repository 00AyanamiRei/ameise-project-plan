import os
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.formatting.rule import CellIsRule

ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data"
OUT_XLSX = ROOT / "AMEISE-Plan.xlsx"

BUDGET_CAP = float(os.getenv("BUDGET_CAP", "225000"))

COLOR_MAP = {
    "CD": "F4A460",  # Code
    "MD": "9ACD32",  # Module Design
    "MN": "40E0D0",  # Manuals
    "SD": "98FB98",  # System Design
    "SP": "F4A1A1",  # Specification
    "C":  "ADD8E6",  # Correcting
    "R":  "87CEFA",  # Review
    "TA": "E6A8D7",  # Acceptance Test
    "TM": "DEB887",  # Module Test
    "TI": "87CEEB",  # Integration Test
    "TS": "D8BFD8",  # System Test
    "RSP": "87CEFA",
    "RSD": "87CEFA",
    "RMD": "87CEFA",
    "RTM": "87CEFA",
    "RTI": "87CEFA",
    "RTS": "87CEFA",
    "CSP": "ADD8E6",
    "CSD": "ADD8E6",
    "CMD": "ADD8E6",
}

def write_df(ws, df, header=True):
    for r in dataframe_to_rows(df, index=False, header=header):
        ws.append(r)
    if header:
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
    thin = Side(style="thin", color="DDDDDD")
    for row in ws.iter_rows():
        for c in row:
            c.border = Border(top=thin, bottom=thin, left=thin, right=thin)

def autosize(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            v = str(cell.value) if cell.value is not None else ""
            max_length = max(max_length, len(v))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 28)

def color_schedule(ws):
    for row in ws.iter_rows(min_row=2, min_col=2):  # skip header and Person col
        for cell in row:
            if not cell.value:
                continue
            token = str(cell.value).split("+")[0].strip()
            color = COLOR_MAP.get(token)
            if color:
                cell.fill = PatternFill("solid", fgColor=color)
            cell.alignment = Alignment(horizontal="center", vertical="center")

def build_schedule_sheet(wb, df):
    ws = wb.create_sheet("Schedule")
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    color_schedule(ws)
    autosize(ws)
    ws.freeze_panes = "B2"

def build_cost_sheet(wb, df):
    ws = wb.create_sheet("Cost Estimation")
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    autosize(ws)
    # Find TOTAL row by label
    total_row = None
    for i in range(2, ws.max_row + 1):
        val = ws.cell(i, 1).value
        if val and str(val).strip().upper() == "TOTAL":
            total_row = i
            break
    if not total_row:
        return
    # Budget Cap (auto)
    ws.cell(row=total_row + 1, column=1).value = "BUDGET_CAP (auto)"
    cap_cell = ws.cell(row=total_row + 1, column=5)
    cap_cell.value = BUDGET_CAP
    cap_cell.number_format = "#,##0.00"
    # Slack = Cap - Total
    ws.cell(row=total_row + 2, column=1).value = "BUDGET_SLACK (auto)"
    slack_cell = ws.cell(row=total_row + 2, column=5)
    slack_cell.value = f"=E{total_row + 1}-E{total_row}"
    slack_cell.number_format = "#,##0.00"
    # Conditional formatting: Total > Cap → red
    ws.conditional_formatting.add(
        f"E{total_row}:E{total_row}",
        CellIsRule(
            operator="greaterThan",
            formula=[f"E{total_row + 1}"],
            fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        ),
    )

def build_optionA_sheet(wb):
    ws = wb.create_sheet("OptionA")
    rows = [
        ["Metric", "Value"],
        ["Person-Months (PM)", 26.36],
        ["Duration (months)", 8.67],
        ["Duration (days)", 263.57],
        ["Avg Developers", 3.04],
        [],
        ["Category","Percent","PM","Months","Days"],
        ["Management",0.07,1.8452,0.6069,18.48],
        ["Specification",0.10,2.6360,0.8670,26.36],
        ["Design",0.27,7.1172,2.3409,71.16],
        ["Coding",0.26,6.8536,2.2542,68.53],
        ["Testing",0.18,4.7448,1.5606,47.06],
        ["Reviews",0.03,0.7908,0.2601,7.91],
        ["Manuals",0.09,2.3724,0.7803,23.53],
        ["TOTAL",1.00,26.36,8.67,263.57],
    ]
    for r in rows:
        ws.append(r)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    autosize(ws)

def main():
    wb = Workbook()
    wb.remove(wb.active)
    df_sched = pd.read_csv(ROOT / "data" / "AMEISE-Schedule-v2-grid.csv")
    df_cost = pd.read_csv(ROOT / "data" / "Cost-Estimation-v2.csv")
    build_schedule_sheet(wb, df_sched)
    build_cost_sheet(wb, df_cost)
    build_optionA_sheet(wb)
    wb.save(OUT_XLSX)
    print(f"Wrote {OUT_XLSX}")

if __name__ == "__main__":
    main()
