import os
import csv
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

# Constants
ROOT = Path(__file__).resolve().parents[1]
TEMPLATE_ENV = os.getenv("TEMPLATE_PATH", "template/AMEISE-planning-template-v1.95-ip.xlsx")
TEMPLATE_PATH = ROOT / TEMPLATE_ENV
OUT_PATH = ROOT / "template" / (Path(TEMPLATE_ENV).stem + "_filled.xlsx")
BUDGET_CAP = float(os.getenv("BUDGET_CAP", "225000"))

# Option A metrics
PM = 26.36
DURATION_MONTHS = 8.67
AVG_DEV = round(PM / DURATION_MONTHS, 2)  # ~3.04
DAYS_PER_MONTH = 30.4

# Effort distribution (fractions)
EFFORT = {
    "Management": 0.07,
    "Specification": 0.10,
    "Design": 0.27,
    "Coding": 0.26,
    "Testing": 0.18,
    "Reviews": 0.03,
    "Manuals": 0.09,
}

# Hourly rates (EUR/h)
RATES = {
    "Axel": 40,
    "Bernd": 40,
    "Christine": 45,
    "Diana": 40,
    "Richard": 50,
    "Stefanie": 40,
    "Thomas": 45,
}

SCHEDULE_CSV = ROOT / "data" / "AMEISE-Schedule-v2-grid.csv"

# ----------------- CSV loading (tolerant to minor format issues) -----------------
def load_schedule():
    target_len = 41  # Person + W1..W40
    with open(SCHEDULE_CSV, newline="", encoding="utf-8") as f:
        rows = list(csv.reader(f))

    if not rows:
        raise ValueError(f"Schedule CSV is empty: {SCHEDULE_CSV}")

    header = ["Person"] + [f"W{i}" for i in range(1, 41)]
    fixed_rows = []
    for idx, row in enumerate(rows[1:], start=2):
        if len(row) < target_len:
            row = row + [""] * (target_len - len(row))
        elif len(row) > target_len:
            tail = row[target_len:]
            if any((c or "").strip() for c in tail):
                raise ValueError(
                    f"Row {idx} has {len(row)} columns (> {target_len}). "
                    f"Extra columns (non-empty) at the end: {tail}. "
                    f"Remove extra separators or quote cells with commas."
                )
            else:
                print(f"[warn] Row {idx} has trailing empty columns ({len(row)}). Trimming to {target_len}.")
                row = row[:target_len]
        fixed_rows.append(row)

    df = pd.DataFrame(fixed_rows, columns=header).fillna("")
    sched = {}
    for _, r in df.iterrows():
        person = str(r["Person"]).strip()
        if not person:
            continue
        weeks = {i: (str(r[f"W{i}"]).strip() if str(r[f"W{i}"]).strip() != "nan" else "") for i in range(1, 41)}
        sched[person] = weeks
    return sched

# ----------------- Helpers for merged cells and label-targeting -----------------
def cell_in_merged_range(ws, r, c):
    for m in ws.merged_cells.ranges:
        if m.min_row <= r <= m.max_row and m.min_col <= c <= m.max_col:
            return m
    return None

def safe_write(ws, r, c, value):
    """
    Write value to cell (r,c), but if it's a merged cell, write to the top-left of its merged range.
    """
    merged = cell_in_merged_range(ws, r, c)
    if merged:
        r, c = merged.min_row, merged.min_col
    ws.cell(r, c, value)

def write_right_of_label(ws, label, value):
    """
    Find the cell with 'label', and write 'value' to the first cell to the right of its merged region.
    """
    pos = find_cell_startswith(ws, label)
    if not pos:
        print(f"[warn] Label not found: {label}")
        return False
    r, c = pos
    target_c = c + 1
    merged = cell_in_merged_range(ws, r, c)
    if merged:
        target_c = merged.max_col + 1
    safe_write(ws, r, target_c, value)
    return True

# ----------------- Generic finders -----------------
def find_cell(ws, text):
    t = str(text).strip().lower()
    for r in range(1, min(ws.max_row, 200)+1):
        for c in range(1, min(ws.max_column, 200)+1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip().lower() == t:
                return r, c
    return None

def find_cell_startswith(ws, text):
    t = str(text).strip().lower()
    for r in range(1, min(ws.max_row, 200)+1):
        for c in range(1, min(ws.max_column, 200)+1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip().lower().startswith(t):
                return r, c
    return None

# ----------------- Week grid + person mapping -----------------
def find_week_grid(ws):
    pos = find_cell(ws, "Week") or find_cell(ws, "week")
    if not pos:
        print("[warn] 'Week' header not found")
        return None
    r, c = pos
    col1 = None
    for cc in range(c+1, min(c+80, ws.max_column+1)):
        v = ws.cell(r, cc).value
        if v in (1, "1"):
            col1 = cc
            break
    if not col1:
        col1 = c+1
    return r, col1

def find_person_row(ws, person, start_row=1, max_row=220, search_cols=8):
    p = person.strip().lower()
    for r in range(start_row, min(max_row, ws.max_row)+1):
        for c in range(1, min(search_cols, ws.max_column)+1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip().lower() == p:
                return r
    return None

def count_assigned_weeks(ws, person_row, col_week1, n_weeks=40):
    used = 0
    for i in range(n_weeks):
        v = ws.cell(person_row, col_week1 + i).value
        if isinstance(v, str) and v.strip():
            used += 1
    return used

# ----------------- Effort Distribution table handling -----------------
def find_effort_headers(ws):
    """
    Locate the Effort Distribution header row with column names:
    'Types of Effort' | '%' | 'PM' | 'months' | 'days'
    Returns (header_row, colmap) where colmap has keys: type, pct, pm, months, days.
    """
    pos = find_cell_startswith(ws, "Types of Effort") or find_cell_startswith(ws, "Types of effort")
    if not pos:
        return None
    r_types, _ = pos

    def scan_row(row):
        colmap = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row, c).value
            if not isinstance(v, str):
                continue
            s = v.strip()
            sl = s.lower()
            if sl == "types of effort" and "type" not in colmap:
                colmap["type"] = c
            elif s == "%" and "pct" not in colmap:
                colmap["pct"] = c
            elif sl == "pm" and "pm" not in colmap:
                colmap["pm"] = c
            elif sl == "months" and "months" not in colmap:
                colmap["months"] = c
            elif sl == "days" and "days" not in colmap:
                colmap["days"] = c
        return colmap

    # Try the found row, then nearby rows (in case headers are just below)
    for rr in (r_types, r_types + 1, r_types - 1):
        if rr < 1 or rr > ws.max_row:
            continue
        colmap = scan_row(rr)
        if all(k in colmap for k in ("type", "pct", "pm", "months", "days")):
            return rr, colmap
    return None

def fill_effort_table(ws):
    headers = find_effort_headers(ws)
    if not headers:
        print("[warn] Effort headers not found; skipping effort table fill")
        return
    header_row, colmap = headers
    start_row = header_row + 1

    i = 0
    for cat, pct in EFFORT.items():
        r = start_row + i
        safe_write(ws, r, colmap["type"], cat)
        safe_write(ws, r, colmap["pct"], round(pct * 100, 2))
        safe_write(ws, r, colmap["pm"], round(PM * pct, 4))
        months = DURATION_MONTHS * pct
        safe_write(ws, r, colmap["months"], round(months, 4))
        safe_write(ws, r, colmap["days"], round(months * DAYS_PER_MONTH, 2))
        i += 1

    # Totals
    r_tot = start_row + len(EFFORT) + 1
    safe_write(ws, r_tot, colmap["type"], "Total Effort")
    safe_write(ws, r_tot, colmap["pm"], round(PM, 2))
    safe_write(ws, r_tot, colmap["months"], round(DURATION_MONTHS, 2))
    safe_write(ws, r_tot, colmap["days"], round(DURATION_MONTHS * DAYS_PER_MONTH, 0))

# ----------------- Schedule grid and Cost table -----------------
def fill_schedule_grid(ws, sched):
    info = find_week_grid(ws)
    if not info:
        print("[warn] Week grid not detected; skip schedule")
        return
    header_row, col_week1 = info
    start_row = header_row + 1
    n_weeks = 40
    for person, weeks in sched.items():
        r = find_person_row(ws, person, start_row=start_row)
        if not r:
            print(f"[info] Person not found in template: {person}")
            continue
        for i in range(1, n_weeks + 1):
            token = weeks.get(i, "")
            safe_write(ws, r, col_week1 + (i - 1), token if token else None)

def fill_cost_table(ws):
    pos = find_cell_startswith(ws, "Cost Estimation")
    if not pos:
        print("[warn] 'Cost Estimation' not found; skip")
        return
    r0, c0 = pos
    name_col = c0
    rate_col = c0 + 1
    weeks_col = c0 + 2
    months_col = c0 + 3
    hours_col = c0 + 4
    total_col = c0 + 5

    info = find_week_grid(ws)
    if not info:
        return
    _, col_week1 = info

    row = r0 + 2
    grand_total = 0.0
    while row <= ws.max_row:
        name = ws.cell(row, name_col).value
        if not name or (isinstance(name, str) and name.strip() == ""):
            break
        name = str(name).strip()
        if name not in RATES:
            row += 1
            continue
        rate = RATES[name]
        person_row = find_person_row(ws, name, start_row=1)
        weeks_used = count_assigned_weeks(ws, person_row, col_week1) if person_row else 0
        hours = weeks_used * 40
        months = round((weeks_used * 7.0) / DAYS_PER_MONTH, 2)
        total = round(hours * rate, 2)

        safe_write(ws, row, rate_col, rate)
        safe_write(ws, row, weeks_col, weeks_used)
        safe_write(ws, row, months_col, months)
        safe_write(ws, row, hours_col, hours)
        safe_write(ws, row, total_col, total)

        grand_total += total
        row += 1

    # Write total if label exists
    for rr in range(row, row + 12):
        v = ws.cell(rr, name_col).value
        if isinstance(v, str) and v.strip().lower().startswith("total project costs"):
            safe_write(ws, rr, total_col, round(grand_total, 2))
            break

# ----------------- Main -----------------
def autodetect_template():
    if TEMPLATE_PATH.exists():
        print(f"[info] Using template from env path: {TEMPLATE_PATH}")
        return TEMPLATE_PATH
    tpl_dir = ROOT / "template"
    if not tpl_dir.exists():
        raise FileNotFoundError(f"Template directory not found: {tpl_dir}")
    candidates = list(tpl_dir.glob("*.xlsx"))
    if not candidates:
        raise FileNotFoundError(f"No .xlsx files found in {tpl_dir}")
    preferred = [c for c in candidates if "planning-template" in c.name.lower() or "ameise" in c.name.lower()]
    chosen = preferred[0] if preferred else candidates[0]
    print(f"[info] TEMPLATE_PATH not found; auto-selected: {chosen}")
    return chosen

def main():
    template_path = autodetect_template()
    sched = load_schedule()

    wb = load_workbook(template_path)
    ws = wb.active

    # Top-left metrics
    write_right_of_label(ws, "Est. Proj. Duration", round(DURATION_MONTHS, 2))
    write_right_of_label(ws, "Est. Avg. Developers", AVG_DEV)
    write_right_of_label(ws, "Est. PersonMonths", round(PM, 2))

    fill_effort_table(ws)
    fill_schedule_grid(ws, sched)
    fill_cost_table(ws)

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"[info] Saved filled workbook to {OUT_PATH}")

if __name__ == "__main__":
    main()
